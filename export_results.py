"""export_results.py - append this run's key numbers to a spreadsheet.

Every run recomputes the same handful of physical observables and prints them
to the terminal, where they scroll past and are lost. This reads the saved
artifacts (data/scan.npz, and data/landscape.npz if it exists), derives the
same numbers analyse.py reports, and appends them as one timestamped row to

    results/run_history.xlsx

so the history accumulates instead of having to be read off by eye each time.
Re-running just adds another row; nothing is overwritten.

If openpyxl is not installed it writes results/run_history.csv instead, which
Excel opens natively, so the step never breaks a run.
"""

import os
from datetime import datetime

import numpy as np

from analyse import (dissociation_limit, fit_minimum, harmonic_frequency,
                     well_depth)

SCAN = "data/scan.npz"
LANDSCAPE = "data/landscape.npz"
OUT_XLSX = "results/run_history.xlsx"
OUT_CSV = "results/run_history.csv"

# (key, spreadsheet header) in the order columns should appear.
COLUMNS = [
    ("timestamp", "timestamp"),
    ("grid_points", "grid points"),
    ("r_eq", "R_e (A)"),
    ("e_min", "E_min (Ha)"),
    ("e_inf", "dissociation limit 2E(H) (Ha)"),
    ("d_e_ev", "well depth D_e (eV)"),
    ("omega", "harmonic freq (cm-1)"),
    ("max_err", "max |VQE-FCI| (mHa)"),
    ("corr_re", "correlation energy at R_e (mHa)"),
    ("iterations", "optimiser iterations"),
    ("theta_move", "theta* drift across scan (rad)"),
]


def collect():
    """Derive the run's observables from the saved .npz artifacts."""
    if not os.path.exists(SCAN):
        raise SystemExit(f"{SCAN} not found - run `python scan.py` first.")

    d = np.load(SCAN)
    r, e_hf, e_fci, e_vqe = d["grid"], d["e_hf"], d["e_fci"], d["e_vqe"]

    r_eq, e_min, coeffs = fit_minimum(r, e_vqe)
    e_inf = dissociation_limit()
    _, d_e_ev = well_depth(e_min, e_inf)
    omega = harmonic_frequency(coeffs)
    max_err = float(np.abs(e_vqe - e_fci).max() * 1000.0)
    corr_re = float((e_hf[int(np.argmin(e_vqe))] - e_min) * 1000.0)
    iterations = int(d["iterations"].sum()) if "iterations" in d.files else None

    # theta* drift is only known once the landscape grid has been built.
    theta_move = None
    if os.path.exists(LANDSCAPE):
        landscape = np.load(LANDSCAPE)
        theta_star = landscape["TH"][landscape["Z"].argmin(axis=0)]
        theta_move = round(float(abs(theta_star[-1] - theta_star[0])), 3)

    return {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "grid_points": int(len(r)),
        "r_eq": round(float(r_eq), 4),
        "e_min": round(float(e_min), 6),
        "e_inf": round(float(e_inf), 6),
        "d_e_ev": round(float(d_e_ev), 4),
        "omega": round(float(omega), 1),
        "max_err": round(max_err, 6),
        "corr_re": round(corr_re, 4),
        "iterations": iterations,
        "theta_move": theta_move,
    }


def append_xlsx(row, out=OUT_XLSX):
    from openpyxl import Workbook, load_workbook

    os.makedirs(os.path.dirname(out), exist_ok=True)
    headers = [header for _, header in COLUMNS]

    if os.path.exists(out):
        wb = load_workbook(out)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "runs"
        ws.append(headers)

    ws.append([row[key] for key, _ in COLUMNS])
    wb.save(out)
    return out


def append_csv(row, out=OUT_CSV):
    import csv

    os.makedirs(os.path.dirname(out), exist_ok=True)
    headers = [header for _, header in COLUMNS]
    new_file = not os.path.exists(out)

    with open(out, "a", newline="") as f:
        writer = csv.writer(f)
        if new_file:
            writer.writerow(headers)
        writer.writerow([row[key] for key, _ in COLUMNS])
    return out


def record(row):
    try:
        return append_xlsx(row)
    except ImportError:
        path = append_csv(row)
        print("  (openpyxl not installed - wrote CSV instead)")
        return path


if __name__ == "__main__":
    data = collect()
    path = record(data)
    print(f"Recorded this run in {path}")
    for key, header in COLUMNS:
        print(f"  {header:32} {data[key]}")
